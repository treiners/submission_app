import json
import os
import mimetypes
import random
import re
import shutil
import hashlib
import importlib.util
import string
import subprocess
import sys
import tempfile
import shlex
from io import BytesIO
from functools import wraps
from pathlib import Path
from urllib import error as urlerror, request as urlrequest

from dotenv import load_dotenv
from flask import (
    Flask, render_template, request, jsonify, session, redirect,
    url_for, send_file, abort, flash
)
from markupsafe import escape
from werkzeug.utils import secure_filename

from src import db, storage, email_util, metadata

load_dotenv()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-key-change-me")

with open(Path(__file__).parent / "config.json") as f:
    SUBMISSION_CONFIG = json.load(f)


def _compute_max_content_length(config):
    """Set a hard upload ceiling above configured per-area limits."""
    areas = config.get("areas", [])
    total_mb = 0
    for area in areas:
        max_size_mb = int(area.get("max_size_mb", 0) or 0)
        max_files = int(area.get("max_files", 1) or 1)
        total_mb += max_size_mb * max_files

    # Keep a safety floor and small multipart overhead headroom.
    hard_ceiling_mb = max(100, total_mb + 20)
    return hard_ceiling_mb * 1024 * 1024


app.config["MAX_CONTENT_LENGTH"] = _compute_max_content_length(SUBMISSION_CONFIG)

PROJECT_ROOT = Path(__file__).parent

db.init_db()

AREAS_BY_KEY = {a["key"]: a for a in SUBMISSION_CONFIG["areas"]}

STORAGE_ENV = {
    "STORAGE_BACKEND": os.environ.get("STORAGE_BACKEND", "local"),
    "LOCAL_UPLOAD_ROOT": os.environ.get("LOCAL_UPLOAD_ROOT", "uploads"),
    "DROPBOX_APP_KEY": os.environ.get("DROPBOX_APP_KEY"),
    "DROPBOX_APP_SECRET": os.environ.get("DROPBOX_APP_SECRET"),
    "DROPBOX_REFRESH_TOKEN": os.environ.get("DROPBOX_REFRESH_TOKEN"),
    "DROPBOX_ROOT_FOLDER": os.environ.get("DROPBOX_ROOT_FOLDER", "/Submissions"),
    "ONEDRIVE_TENANT_ID": os.environ.get("ONEDRIVE_TENANT_ID"),
    "ONEDRIVE_CLIENT_ID": os.environ.get("ONEDRIVE_CLIENT_ID"),
    "ONEDRIVE_CLIENT_SECRET": os.environ.get("ONEDRIVE_CLIENT_SECRET"),
    "ONEDRIVE_DRIVE_ID": os.environ.get("ONEDRIVE_DRIVE_ID"),
    "ONEDRIVE_ROOT_FOLDER": os.environ.get("ONEDRIVE_ROOT_FOLDER", "/Submissions"),
}

CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ"  # letters only, excludes I/O to reduce confusion
CODE_PATTERN = re.compile(r"^[A-Z]{6}$")
OLLAMA_ENDPOINT = os.environ.get("OLLAMA_ENDPOINT", "http://127.0.0.1:11434")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")
MARKING_TEXT_PREVIEW_EXTENSIONS = {".mod", ".run", ".dat"}
MARKING_TEXT_PREVIEW_MAX_BYTES = int(os.environ.get("MARKING_TEXT_PREVIEW_MAX_BYTES", str(2 * 1024 * 1024)))
MARKING_DOCUMENT_PREVIEW_EXTENSIONS = {".pdf", ".docx"}
MARKING_VIDEO_PREVIEW_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v", ".wmv"}
MARKING_DOCX_PREVIEW_MAX_PARAGRAPHS = int(os.environ.get("MARKING_DOCX_PREVIEW_MAX_PARAGRAPHS", "900"))
MARKING_DOCX_PDF_PREVIEW_DIR = "marking_pdf_previews"
MARKING_DOCX_PDF_CONVERTER = str(
    os.environ.get(
        "MARKING_DOCX_PDF_CONVERTER",
        SUBMISSION_CONFIG.get("marking_docx_pdf_converter", "auto"),
    )
).strip().lower()
ACTIVE_TEMPLATE_DOCX_PATH = PROJECT_ROOT / "marking_template" / "active_template.docx"
ACTIVE_TEMPLATE_JSON_PATH = PROJECT_ROOT / "marking_template" / "test_case" / "active_template.json"


def _normalize_question_id(value):
    raw = str(value or "").strip().upper()
    if not raw:
        return None

    if raw.startswith("Q"):
        digits = raw[1:]
    else:
        digits = raw

    if not digits.isdigit():
        return None

    number = int(digits)
    if number <= 0:
        return None
    return f"Q{number}"


def _normalize_area_question_targets(area):
    """Return None for all-questions scope, else a set of normalized question IDs."""
    if "questions" in area:
        raw_value = area.get("questions")
    elif "question" in area:
        raw_value = area.get("question")
    else:
        return None

    if raw_value in (None, "", []):
        return None

    values = raw_value if isinstance(raw_value, list) else [raw_value]
    normalized = set()
    for item in values:
        question_id = _normalize_question_id(item)
        if not question_id:
            raise ValueError(
                f"Invalid question mapping '{item}' for area '{area.get('key', '?')}'. "
                "Use values like 2, '2', or 'Q2'."
            )
        normalized.add(question_id)

    if not normalized:
        return None
    return normalized


def _build_area_question_map(config):
    mapping = {}
    for area in config.get("areas", []):
        area_key = str(area.get("key", "")).strip()
        if not area_key:
            continue
        mapping[area_key] = _normalize_area_question_targets(area)
    return mapping


AREA_QUESTION_MAP = _build_area_question_map(SUBMISSION_CONFIG)


def _question_sort_key(question_id):
    parts = re.split(r"(\d+)", str(question_id or ""))
    normalized = []
    for part in parts:
        if part.isdigit():
            normalized.append((0, int(part)))
        else:
            normalized.append((1, part.lower()))
    return normalized


def _format_marks_label(max_score):
    if max_score in (None, ""):
        return ""
    try:
        numeric = float(max_score)
    except (TypeError, ValueError):
        return ""
    if numeric.is_integer():
        return f"{int(numeric)}M"
    return f"{numeric}M"


def _load_active_template_questions():
    context = {
        "status": "missing",
        "warning": "",
        "questions": [],
        "question_ids": [],
        "source_docx_exists": ACTIVE_TEMPLATE_DOCX_PATH.exists(),
        "source_json_exists": ACTIVE_TEMPLATE_JSON_PATH.exists(),
    }

    if not ACTIVE_TEMPLATE_JSON_PATH.exists():
        context["warning"] = (
            "Active template JSON is missing. Marking is using extracted questions as a fallback."
        )
        return context

    try:
        payload = json.loads(ACTIVE_TEMPLATE_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        context["status"] = "invalid"
        context["warning"] = (
            f"Could not read active template JSON: {error}. "
            "Marking is using extracted questions as a fallback."
        )
        return context

    raw_questions = payload.get("questions")
    if not isinstance(raw_questions, list):
        context["status"] = "invalid"
        context["warning"] = (
            "Active template JSON is invalid (missing 'questions' list). "
            "Marking is using extracted questions as a fallback."
        )
        return context

    normalized = []
    seen = set()
    errors = []

    for index, item in enumerate(raw_questions, start=1):
        if not isinstance(item, dict):
            errors.append(f"Entry {index} is not an object.")
            continue

        question_id = _normalize_question_id(item.get("question_id"))
        if not question_id:
            errors.append(f"Entry {index} has an invalid question_id.")
            continue
        if question_id in seen:
            errors.append(f"Duplicate question_id '{question_id}' in active template JSON.")
            continue

        question_prompt = str(item.get("question_prompt", "") or "").strip()
        max_score_raw = item.get("max_score")
        max_score = None
        if max_score_raw not in (None, ""):
            try:
                max_score = float(max_score_raw)
            except (TypeError, ValueError):
                errors.append(f"Question {question_id} has invalid max_score '{max_score_raw}'.")

        normalized.append({
            "question_id": question_id,
            "question_prompt": question_prompt,
            "max_score": max_score,
            "marks_label": _format_marks_label(max_score),
        })
        seen.add(question_id)

    if not normalized:
        context["status"] = "invalid"
        context["warning"] = (
            "Active template JSON has no valid questions. "
            "Marking is using extracted questions as a fallback."
        )
        return context

    normalized = sorted(normalized, key=lambda item: _question_sort_key(item["question_id"]))
    context["questions"] = normalized
    context["question_ids"] = [item["question_id"] for item in normalized]

    if errors:
        context["status"] = "partial"
        context["warning"] = "Active template loaded with warnings: " + " ".join(errors)
    else:
        context["status"] = "active"

    return context


def _resolve_docx_pdf_converter(preferred=None):
    converter = str(preferred or MARKING_DOCX_PDF_CONVERTER or "auto").strip().lower()
    if converter not in {"auto", "libreoffice", "docx2pdf", "dxpdf"}:
        converter = "auto"

    if converter == "auto":
        if shutil.which("soffice") or shutil.which("libreoffice"):
            return "libreoffice"
        if importlib.util.find_spec("dxpdf") is not None:
            return "dxpdf"
        return "docx2pdf"

    return converter


def generate_code():
    for _ in range(50):
        code = "".join(random.choices(CODE_ALPHABET, k=6))
        if not db.code_exists(code):
            return code
    raise RuntimeError("Could not generate a unique code, database may be full of codes.")


def get_client_ip():
    # Respect X-Forwarded-For if the app sits behind a reverse proxy (Nginx),
    # otherwise fall back to the direct connection IP.
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr


def validate_file(area, file_storage):
    ext = Path(file_storage.filename).suffix.lower()
    if ext not in area["allowed_extensions"]:
        return f"'{file_storage.filename}' has an invalid file type for {area['label']} (allowed: {', '.join(area['allowed_extensions'])})."

    file_storage.stream.seek(0, os.SEEK_END)
    size_bytes = file_storage.stream.tell()
    file_storage.stream.seek(0)
    max_bytes = area["max_size_mb"] * 1024 * 1024
    if size_bytes > max_bytes:
        return f"'{file_storage.filename}' exceeds the {area['max_size_mb']}MB limit for {area['label']}."
    if size_bytes == 0:
        return f"'{file_storage.filename}' is empty."
    return None


def _submission_metadata_folder(submission_record):
    local_files = [Path(file["storage_location"]) for file in submission_record.get("files", [])]
    if not local_files:
        return None
    return local_files[0].parents[1] / "metadata"


def _submission_root_folder(submission_record):
    local_files = [Path(file["storage_location"]) for file in submission_record.get("files", [])]
    if not local_files:
        return None
    return local_files[0].parents[1]


def _area_applies_to_question(area_key, question_id):
    scope = AREA_QUESTION_MAP.get(area_key)
    if not scope:
        return True
    normalized_question = _normalize_question_id(question_id)
    if not normalized_question:
        return True
    return normalized_question in scope


def _collect_marking_text_files(submission_record, question_id=None):
    submission_root = _submission_root_folder(submission_record)
    if not submission_root:
        return []

    try:
        resolved_root = submission_root.resolve()
    except OSError:
        return []

    entries = []
    seen_paths = set()
    for file_record in submission_record.get("files", []):
        area_key = file_record.get("area_key")
        if question_id and not _area_applies_to_question(area_key, question_id):
            continue

        path = Path(file_record.get("storage_location", ""))
        if path.suffix.lower() not in MARKING_TEXT_PREVIEW_EXTENSIONS:
            continue
        if not path.exists() or not path.is_file():
            continue

        try:
            resolved_path = path.resolve()
            relative_path = resolved_path.relative_to(resolved_root).as_posix()
        except (OSError, ValueError):
            continue

        if relative_path in seen_paths:
            continue
        seen_paths.add(relative_path)

        original_name = file_record.get("original_filename") or path.name
        area_label = file_record.get("area_label") or file_record.get("area_key") or "file"
        entries.append({
            "path": relative_path,
            "label": f"{area_label}: {original_name}",
        })

    return sorted(entries, key=lambda item: item["path"].lower())


def _collect_marking_document_files(submission_record, question_id=None):
    submission_root = _submission_root_folder(submission_record)
    if not submission_root:
        return []

    try:
        resolved_root = submission_root.resolve()
    except OSError:
        return []

    entries = []
    seen_paths = set()
    for file_record in submission_record.get("files", []):
        area_key = file_record.get("area_key")
        if question_id and not _area_applies_to_question(area_key, question_id):
            continue

        path = Path(file_record.get("storage_location", ""))
        extension = path.suffix.lower()
        if extension not in MARKING_DOCUMENT_PREVIEW_EXTENSIONS and extension not in MARKING_VIDEO_PREVIEW_EXTENSIONS:
            continue
        if not path.exists() or not path.is_file():
            continue

        try:
            resolved_path = path.resolve()
            relative_path = resolved_path.relative_to(resolved_root).as_posix()
        except (OSError, ValueError):
            continue

        if relative_path in seen_paths:
            continue
        seen_paths.add(relative_path)

        original_name = file_record.get("original_filename") or path.name
        area_label = file_record.get("area_label") or file_record.get("area_key") or "file"
        entries.append({
            "path": relative_path,
            "label": f"{area_label}: {original_name}",
            "extension": extension,
            "name": original_name,
            "media_type": "video" if extension in MARKING_VIDEO_PREVIEW_EXTENSIONS else "document",
        })

    return sorted(entries, key=lambda item: item["path"].lower())


def _resolve_submission_relative_file(submission_record, relative_path):
    submission_root = _submission_root_folder(submission_record)
    if not submission_root:
        return None

    try:
        resolved_root = submission_root.resolve()
        candidate = (resolved_root / relative_path).resolve()
        candidate.relative_to(resolved_root)
    except (OSError, ValueError):
        return None

    if not candidate.exists() or not candidate.is_file():
        return None
    return candidate


def _docx_pdf_preview_output_path(submission_record, relative_path, source_path):
    metadata_folder = _submission_metadata_folder(submission_record)
    if not metadata_folder:
        return None

    preview_dir = metadata_folder / MARKING_DOCX_PDF_PREVIEW_DIR
    preview_dir.mkdir(parents=True, exist_ok=True)

    source_stat = source_path.stat()
    source_signature = f"{relative_path}|{int(source_stat.st_mtime)}|{source_stat.st_size}"
    digest = hashlib.sha1(source_signature.encode("utf-8")).hexdigest()[:12]
    base_name = secure_filename(Path(relative_path).stem) or "document"
    return preview_dir / f"{base_name}_{digest}.pdf"


def _convert_docx_to_pdf_with_docx2pdf(source_docx, target_pdf):
    try:
        from docx2pdf import convert as docx2pdf_convert
    except ImportError as error:
        raise RuntimeError(
            "docx2pdf is required for DOCX-to-PDF preview conversion. "
            "Install dependencies from requirements.txt."
        ) from error

    try:
        docx2pdf_convert(str(source_docx.resolve()), str(target_pdf.resolve()))
    except Exception as error:
        raise RuntimeError(
            "docx2pdf conversion failed. On macOS this usually requires Microsoft Word to be installed and available. "
            f"({error})"
        ) from error

    if not target_pdf.exists() or target_pdf.stat().st_size == 0:
        raise RuntimeError("docx2pdf did not produce a valid PDF output.")


def _convert_docx_to_pdf_with_dxpdf(source_docx, target_pdf):
    try:
        import dxpdf
    except ImportError as error:
        raise RuntimeError(
            "dxpdf is required for DOCX-to-PDF preview conversion. "
            "Install dependencies from requirements.txt."
        ) from error

    try:
        dxpdf.convert_file(str(source_docx.resolve()), str(target_pdf.resolve()))
    except Exception as error:
        raise RuntimeError(f"dxpdf conversion failed: {error}") from error

    if not target_pdf.exists() or target_pdf.stat().st_size == 0:
        raise RuntimeError("dxpdf did not produce a valid PDF output.")


def _convert_docx_to_pdf_with_libreoffice(source_docx, target_pdf):
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice is not installed or not available on PATH.")

    output_dir = target_pdf.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    command = [
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(source_docx),
    ]
    completed = subprocess.run(command, capture_output=True, text=True)
    if completed.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed: {completed.stderr.strip() or completed.stdout.strip() or 'unknown error'}"
        )

    produced_pdf = output_dir / f"{source_docx.stem}.pdf"
    if produced_pdf.exists() and produced_pdf != target_pdf:
        if target_pdf.exists():
            try:
                target_pdf.unlink()
            except OSError:
                pass
        produced_pdf.replace(target_pdf)

    if not target_pdf.exists() or target_pdf.stat().st_size == 0:
        raise RuntimeError("LibreOffice did not produce a valid PDF output.")


def _convert_docx_to_pdf(source_docx, target_pdf, converter=None):
    resolved = _resolve_docx_pdf_converter(converter)
    if resolved == "libreoffice":
        return _convert_docx_to_pdf_with_libreoffice(source_docx, target_pdf)
    if resolved == "docx2pdf":
        return _convert_docx_to_pdf_with_docx2pdf(source_docx, target_pdf)
    if resolved == "dxpdf":
        return _convert_docx_to_pdf_with_dxpdf(source_docx, target_pdf)
    raise RuntimeError(f"Unsupported DOCX-to-PDF converter: {resolved}")


def _ensure_docx_pdf_preview(submission_record, relative_path, source_path):
    output_path = _docx_pdf_preview_output_path(submission_record, relative_path, source_path)
    if not output_path:
        raise RuntimeError("Could not determine metadata folder for PDF preview output.")

    if output_path.exists():
        return output_path

    temp_output = output_path.with_suffix(".tmp.pdf")
    if temp_output.exists():
        try:
            temp_output.unlink()
        except OSError:
            pass

    _convert_docx_to_pdf(source_path, temp_output)
    temp_output.replace(output_path)
    return output_path


def _find_submission_run_path(submission_record, run_filename):
    candidates = []
    for file_record in submission_record.get("files", []):
        path = Path(file_record.get("storage_location", ""))
        if path.suffix.lower() != ".run":
            continue
        if path.name != run_filename:
            continue
        if path.exists() and path.is_file():
            candidates.append(path)
    if not candidates:
        return None
    return sorted(candidates, key=lambda p: str(p))[0]


def _parse_ampl_run_references(run_path):
    try:
        content = run_path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return []

    # Strip inline comments first, then parse statement-by-statement.
    without_comments = "\n".join(line.split("#", 1)[0] for line in content.splitlines())
    references = []
    seen = set()
    for statement in without_comments.split(";"):
        match = re.match(r"^\s*(model|data)\b(.*)$", statement, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            continue

        tail = (match.group(2) or "").strip()
        if not tail:
            continue

        try:
            tokens = shlex.split(tail)
        except ValueError:
            tokens = tail.split()

        for token in tokens:
            token = token.strip().rstrip(",")
            if not token:
                continue
            lowered = token.lower()
            if lowered in {"model", "data"}:
                continue
            if lowered.endswith(":"):
                continue
            if token in seen:
                continue
            seen.add(token)
            references.append(token)

    return references


def _collect_analysis_text_files_for_run(submission_record, run_record):
    run_filename = run_record.get("run_filename", "")
    run_path = _find_submission_run_path(submission_record, run_filename)
    submission_root = _submission_root_folder(submission_record)
    if not run_path or not submission_root:
        return []

    try:
        resolved_root = submission_root.resolve()
        resolved_run_path = run_path.resolve()
    except OSError:
        return []

    files = []
    seen_relative = set()

    def add_candidate(path_obj, source_label):
        try:
            resolved = path_obj.resolve()
            relative_path = resolved.relative_to(resolved_root).as_posix()
        except (OSError, ValueError):
            return
        if not resolved.exists() or not resolved.is_file():
            return
        if relative_path in seen_relative:
            return
        seen_relative.add(relative_path)
        files.append({
            "path": relative_path,
            "name": resolved.name,
            "label": source_label,
        })

    add_candidate(resolved_run_path, "Run script")

    for ref in _parse_ampl_run_references(resolved_run_path):
        ref_path = Path(ref)
        if ref_path.is_absolute():
            continue
        add_candidate((resolved_run_path.parent / ref_path), "Referenced by run")

    return sorted(files, key=lambda item: item["path"].lower())


def _load_marking_previews(submission_record):
    previews = []
    metadata_folder = _submission_metadata_folder(submission_record)
    if not metadata_folder:
        return previews

    for preview_path in sorted(metadata_folder.glob("*.answers.json")):
        try:
            preview_data = json.loads(preview_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        previews.append({
            "file": preview_path.name,
            "data": preview_data,
        })
    return previews


def _extract_marking_preview_for_submission_file(submission_record, file_record):
    metadata_folder = _submission_metadata_folder(submission_record)
    if not metadata_folder:
        return False

    docx_path = Path(file_record["storage_location"])
    if not docx_path.exists() or docx_path.suffix.lower() != ".docx":
        return False

    preview = metadata.extract_marking_preview(
        submission_docx=docx_path,
        submission_config=SUBMISSION_CONFIG,
        project_root=PROJECT_ROOT,
    )
    out_file = metadata_folder / f"{file_record['area_key']}_{file_record['stored_filename']}.answers.json"
    metadata.write_metadata(out_file, preview)
    return True


def _extract_json_object(text):
    text = (text or "").strip()
    if not text:
        return None

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    start = text.find("{")
    if start == -1:
        return None

    depth = 0
    in_string = False
    escaped = False
    for idx in range(start, len(text)):
        char = text[idx]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                candidate = text[start:idx + 1]
                try:
                    parsed = json.loads(candidate)
                    return parsed if isinstance(parsed, dict) else None
                except json.JSONDecodeError:
                    return None
    return None


def _parse_max_score_from_marks_label(marks_label):
    match = re.search(r"(\d+)", marks_label or "")
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return 5.0
    return 5.0


def _build_marking_ai_prompt(question_id, prompt, answer, marks_label):
    max_score = _parse_max_score_from_marks_label(marks_label)
    return (
        "You are helping an academic marker draft a provisional mark. "
        "Return strict JSON only and no markdown. "
        "Required keys: score, feedback_comment, rationale, minimum_requirements_met, strengths, gaps. "
        "score must be numeric in [0, max_score]. "
        "minimum_requirements_met must be true or false. "
        "strengths and gaps must be short string arrays.\n\n"
        f"Question ID: {question_id}\n"
        f"Question prompt: {prompt}\n"
        f"Student answer: {answer}\n"
        f"Mark label: {marks_label or 'n/a'}\n"
        f"max_score: {max_score}\n"
    )


def _generate_ai_marking_suggestion(question_id, prompt, answer, marks_label):
    max_score = _parse_max_score_from_marks_label(marks_label)
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": _build_marking_ai_prompt(question_id, prompt, answer, marks_label),
        "stream": False,
        "options": {
            "temperature": 0.1,
        },
    }

    endpoint = OLLAMA_ENDPOINT.rstrip("/") + "/api/generate"
    req = urlrequest.Request(
        endpoint,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    timeout_seconds = int(os.environ.get("OLLAMA_TIMEOUT_SECONDS", "90"))

    with urlrequest.urlopen(req, timeout=timeout_seconds) as resp:
        body = json.loads(resp.read().decode("utf-8"))

    raw_text = str(body.get("response", "")).strip()
    parsed = _extract_json_object(raw_text)
    if not parsed:
        return {
            "ok": False,
            "error": "Model output was not valid JSON.",
            "raw_response": raw_text,
        }

    score = parsed.get("score")
    try:
        numeric_score = float(score)
    except (TypeError, ValueError):
        numeric_score = None

    if numeric_score is not None:
        numeric_score = max(0.0, min(max_score, numeric_score))

    suggestion = {
        "score": numeric_score,
        "feedback_comment": str(parsed.get("feedback_comment", "")).strip(),
        "rationale": str(parsed.get("rationale", "")).strip(),
        "minimum_requirements_met": bool(parsed.get("minimum_requirements_met", False)),
        "strengths": parsed.get("strengths", []) if isinstance(parsed.get("strengths", []), list) else [],
        "gaps": parsed.get("gaps", []) if isinstance(parsed.get("gaps", []), list) else [],
        "max_score": max_score,
        "model": OLLAMA_MODEL,
    }
    return {"ok": True, "suggestion": suggestion, "raw_response": raw_text}


@app.context_processor
def inject_template_globals():
    return {
        "form_version": SUBMISSION_CONFIG.get("form_version"),
    }


# ---------- Public submission routes ----------

@app.route("/")
def index():
    return render_template(
        "index.html",
        assignment_title=SUBMISSION_CONFIG["assignment_title"],
        areas=SUBMISSION_CONFIG["areas"],
    )


@app.route("/api/config")
def api_config():
    return jsonify(SUBMISSION_CONFIG)


@app.route("/api/submission-preview", methods=["POST"])
def api_submission_preview():
    payload = request.get_json(silent=True) or {}
    code = str(payload.get("code", "")).strip().upper()

    if not CODE_PATTERN.fullmatch(code):
        return jsonify({
            "ok": False,
            "error": "Please enter a valid 6-letter confirmation code.",
        }), 400

    preview = db.get_submission_preview_by_code(code)
    if not preview:
        return jsonify({
            "ok": False,
            "error": "No submission was found for that code.",
        }), 404

    return jsonify({"ok": True, "submission": preview})


@app.route("/api/submit", methods=["POST"])
def api_submit():
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    legacy_name = request.form.get("name", "").strip()

    # Backward-compatible fallback for old clients still posting single name.
    if not first_name and not last_name and legacy_name:
        parts = legacy_name.split(None, 1)
        first_name = parts[0]
        last_name = parts[1] if len(parts) > 1 else ""

    name = " ".join(part for part in [first_name, last_name] if part)
    student_id = request.form.get("student_id", "").strip()
    email = request.form.get("email", "").strip()

    errors = []
    if not first_name:
        errors.append("First name is required.")
    if not last_name:
        errors.append("Last name is required.")
    if not student_id:
        errors.append("Student ID is required.")
    if not email:
        errors.append("Email is required for the confirmation email.")

    # Validate every configured area: satisfied by either an attached file, or
    # the "I declare that I did not upload this file" checkbox.
    files_by_area = {}
    declared_areas = []
    for area in SUBMISSION_CONFIG["areas"]:
        key = area["key"]
        declared = request.form.get(f"{key}_declared") == "true"
        uploaded = request.files.getlist(key)
        uploaded = [f for f in uploaded if f and f.filename]

        if declared:
            if uploaded:
                errors.append(
                    f"You attached a file for '{area['label']}' but also marked it as "
                    f"not submitted. Please remove the file or untick the box."
                )
                continue
            declared_areas.append(area)
            continue

        if len(uploaded) == 0:
            errors.append(
                f"Please attach a file for '{area['label']}', or tick the box "
                f"confirming you did not submit it."
            )
            continue
        if len(uploaded) > area["max_files"]:
            errors.append(f"Only {area['max_files']} file(s) allowed for '{area['label']}'.")
            continue

        for f in uploaded:
            err = validate_file(area, f)
            if err:
                errors.append(err)

        files_by_area[key] = uploaded

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    # Everything valid - persist to a temp dir, then hand off to storage backend
    code = generate_code()
    dest_folder = f"{student_id}_{secure_filename(name)}_{code}"
    primary_backend, fallback_backend = storage.get_storage_backend(STORAGE_ENV)

    submission_id = db.create_submission(
        name=name, student_id=student_id, code=code,
        ip_address=get_client_ip(), user_agent=request.headers.get("User-Agent", ""),
        email=email, storage_backend=primary_backend.name,
        status="processing",
    )

    saved_filenames = []
    used_fallback = False
    extraction_areas = set(SUBMISSION_CONFIG.get("marking_extraction_areas", ["report"]))

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            for area_key, files in files_by_area.items():
                area = AREAS_BY_KEY[area_key]
                for f in files:
                    safe_name = secure_filename(f.filename)
                    tmp_path = Path(tmpdir) / safe_name
                    f.save(tmp_path)

                    area_folder = f"{dest_folder}/{area_key}"
                    try:
                        result = primary_backend.upload_file(tmp_path, area_folder, safe_name)
                    except storage.StorageError:
                        result = fallback_backend.upload_file(tmp_path, area_folder, safe_name)
                        used_fallback = True

                    file_metadata = metadata.extract_file_metadata(
                        tmp_path, f.filename, area_key
                    )
                    area_scope = AREA_QUESTION_MAP.get(area_key)
                    file_metadata["question_scope"] = sorted(area_scope) if area_scope else "all"
                    metadata_path = (
                        Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"])
                        / dest_folder
                        / "metadata"
                        / f"{area_key}_{safe_name}.json"
                    )
                    metadata.write_metadata(metadata_path, file_metadata)

                    if area_key in extraction_areas and tmp_path.suffix.lower() == ".docx":
                        try:
                            marking_preview = metadata.extract_marking_preview(
                                submission_docx=tmp_path,
                                submission_config=SUBMISSION_CONFIG,
                                project_root=PROJECT_ROOT,
                            )
                            preview_path = (
                                Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"])
                                / dest_folder
                                / "metadata"
                                / f"{area_key}_{safe_name}.answers.json"
                            )
                            metadata.write_metadata(preview_path, marking_preview)
                        except (OSError, metadata.zipfile.BadZipFile, metadata.ElementTree.ParseError) as error:
                            app.logger.warning(
                                "Marking preview extraction failed for submission %s (%s): %s",
                                submission_id,
                                safe_name,
                                error,
                            )

                    db.add_submission_file(
                        submission_id=submission_id, area_key=area_key, area_label=area["label"],
                        original_filename=f.filename, stored_filename=safe_name,
                        storage_location=result["location"], size_bytes=tmp_path.stat().st_size,
                    )
                    saved_filenames.append(f.filename)

                    if area_key == "ampl_code" and tmp_path.suffix.lower() == ".zip":
                        extract_root = Path(tmpdir) / "ampl_extracted"
                        try:
                            extracted_files = metadata.extract_ampl_archive(tmp_path, extract_root)
                        except (metadata.zipfile.BadZipFile, OSError) as error:
                            app.logger.warning("AMPL archive extraction failed for submission %s: %s", submission_id, error)
                            extracted_files = []

                        for extracted_path, archive_name in extracted_files:
                            relative_path = extracted_path.relative_to(extract_root)
                            extracted_folder = f"{area_folder}/{relative_path.parent.as_posix()}"
                            extracted_name = extracted_path.name
                            try:
                                extracted_result = primary_backend.upload_file(
                                    extracted_path, extracted_folder, extracted_name
                                )
                            except storage.StorageError:
                                extracted_result = fallback_backend.upload_file(
                                    extracted_path, extracted_folder, extracted_name
                                )
                                used_fallback = True

                            extracted_metadata = metadata.extract_file_metadata(
                                extracted_path, archive_name, area_key
                            )
                            extracted_metadata["extracted_from"] = f.filename
                            extracted_metadata["question_scope"] = sorted(area_scope) if area_scope else "all"
                            extracted_metadata_path = (
                                Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"])
                                / dest_folder
                                / "metadata"
                                / f"{area_key}_{relative_path.as_posix().replace('/', '_')}.json"
                            )
                            metadata.write_metadata(extracted_metadata_path, extracted_metadata)
                            db.add_submission_file(
                                submission_id=submission_id, area_key=area_key, area_label=area["label"],
                                original_filename=archive_name, stored_filename=extracted_name,
                                storage_location=extracted_result["location"],
                                size_bytes=extracted_path.stat().st_size,
                            )

        for area in declared_areas:
            db.add_declaration(submission_id, area["key"], area["label"])

        if used_fallback:
            conn = db.get_connection()
            conn.execute(
                "UPDATE submissions SET storage_backend = ?, storage_note = ? WHERE id = ?",
                (fallback_backend.name, "Primary storage backend failed; saved locally instead.", submission_id),
            )
            conn.commit()
            conn.close()

        email_sent = False
        try:
            email_util.send_confirmation_email(
                gmail_address=os.environ.get("GMAIL_ADDRESS"),
                gmail_app_password=os.environ.get("GMAIL_APP_PASSWORD"),
                from_name=os.environ.get("EMAIL_FROM_NAME", "Assignment Submission System"),
                to_address=email, name=name, student_id=student_id, code=code,
                filenames=saved_filenames, assignment_title=SUBMISSION_CONFIG["assignment_title"],
                declared_labels=[a["label"] for a in declared_areas],
            )
            email_sent = True
            db.mark_email_sent(submission_id, True)
        except Exception as e:
            app.logger.warning(f"Email send failed for submission {submission_id}: {e}")

        db.update_submission_status(submission_id, "completed")

        return jsonify({
            "ok": True,
            "code": code,
            "email_sent": email_sent,
            "storage_note": "Saved locally (cloud storage unavailable)." if used_fallback else None,
        })
    except Exception as error:
        failure_reason = f"{type(error).__name__}: {error}"
        app.logger.exception("Submission processing failed for submission %s", submission_id)
        db.update_submission_status(submission_id, "unprocessed", failure_reason)
        return jsonify({
            "ok": True,
            "email_sent": False,
            "code": code,
            "submission_id": submission_id,
            "processing_note": (
                "Your files were received, but automatic processing could not be completed. "
                "Staff can still access your submission and process it manually."
            ),
            "failure_reason": failure_reason[:200],
        }), 200


# ---------- Admin auth ----------

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if (username == os.environ.get("ADMIN_USERNAME") and
                password == os.environ.get("ADMIN_PASSWORD") and
                os.environ.get("ADMIN_PASSWORD")):
            session["admin_logged_in"] = True
            next_url = request.args.get("next") or url_for("admin_dashboard")
            return redirect(next_url)
        flash("Invalid username or password.")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


# ---------- Admin dashboard ----------

@app.route("/admin")
@login_required
def admin_dashboard():
    search = request.args.get("q", "").strip()
    marking_filter = request.args.get("marking", "all")
    sort_key = request.args.get("sort", "submitted_at")
    sort_dir = request.args.get("dir", "desc")
    return _render_admin_dashboard(
        search=search,
        marking_filter=marking_filter,
        sort_key=sort_key,
        sort_dir=sort_dir,
    )


def _build_dashboard_context(search="", marking_filter="all", sort_key="submitted_at", sort_dir="desc"):
    marking_filter = str(marking_filter or "all").strip().lower()
    if marking_filter not in {"all", "included", "excluded"}:
        marking_filter = "all"

    sort_key = str(sort_key or "submitted_at").strip().lower()
    if sort_key not in {
        "submitted_at",
        "name",
        "student_id",
        "code",
        "status",
        "marking",
        "storage_backend",
        "email_sent",
        "ip_address",
    }:
        sort_key = "submitted_at"

    sort_dir = str(sort_dir or "desc").strip().lower()
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc"

    submissions = db.list_submissions(search=search or None, marking_filter=marking_filter)

    status_order = {
        "completed": 0,
        "processing": 1,
        "unprocessed": 2,
        "failed": 3,
    }

    def dashboard_sort_value(submission):
        if sort_key == "submitted_at":
            return submission.get("submitted_at") or ""
        if sort_key == "name":
            return (submission.get("name") or "").lower()
        if sort_key == "student_id":
            return (submission.get("student_id") or "").lower()
        if sort_key == "code":
            return (submission.get("code") or "").lower()
        if sort_key == "status":
            return (status_order.get(str(submission.get("status") or "").lower(), 99), submission.get("submitted_at") or "")
        if sort_key == "marking":
            return (1 if submission.get("marking_excluded") else 0, submission.get("submitted_at") or "")
        if sort_key == "storage_backend":
            return (submission.get("storage_backend") or "").lower()
        if sort_key == "email_sent":
            return (1 if submission.get("email_sent") else 0, submission.get("submitted_at") or "")
        if sort_key == "ip_address":
            return (submission.get("ip_address") or "").lower()
        return submission.get("submitted_at") or ""

    submissions = sorted(submissions, key=dashboard_sort_value, reverse=(sort_dir == "desc"))

    included_count = len(db.list_submissions(marking_filter="included"))
    excluded_count = len(db.list_submissions(marking_filter="excluded"))

    active_template_context = _load_active_template_questions()

    return {
        "submissions": submissions,
        "search": search,
        "marking_filter": marking_filter,
        "sort_key": sort_key,
        "sort_dir": sort_dir,
        "included_count": included_count,
        "excluded_count": excluded_count,
        "assignment_title": SUBMISSION_CONFIG["assignment_title"],
        "form_version": SUBMISSION_CONFIG.get("form_version"),
        "active_template_docx_exists": ACTIVE_TEMPLATE_DOCX_PATH.exists(),
        "active_template_json_exists": ACTIVE_TEMPLATE_JSON_PATH.exists(),
        "active_template_docx_name": ACTIVE_TEMPLATE_DOCX_PATH.name,
        "active_template_json_name": ACTIVE_TEMPLATE_JSON_PATH.name,
        "active_template_question_count": len(active_template_context["question_ids"]),
        "active_template_status": active_template_context["status"],
        "active_template_warning": active_template_context["warning"],
    }


def _render_admin_dashboard(search="", marking_filter="all", sort_key="submitted_at", sort_dir="desc", template_upload_errors=None, status_code=200):
    context = _build_dashboard_context(
        search=search,
        marking_filter=marking_filter,
        sort_key=sort_key,
        sort_dir=sort_dir,
    )
    context["template_upload_errors"] = template_upload_errors or []
    return render_template("admin_dashboard.html", **context), status_code


@app.route("/admin/template/upload", methods=["POST"])
@login_required
def admin_template_upload():
    uploaded = request.files.get("template_docx")
    errors = []

    if not uploaded or not uploaded.filename:
        errors.append("Please choose a DOCX template file to upload.")
    else:
        extension = Path(uploaded.filename).suffix.lower()
        if extension != ".docx":
            errors.append("Template upload only accepts .docx files.")

    if errors:
        return _render_admin_dashboard(template_upload_errors=errors, status_code=400)

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            candidate = Path(tmpdir) / (secure_filename(uploaded.filename) or "template.docx")
            uploaded.save(candidate)

            template_payload = metadata.build_active_template_case_file(candidate)
            marker_errors = template_payload.get("errors", [])
            if marker_errors:
                return _render_admin_dashboard(
                    template_upload_errors=["Template validation failed:", *marker_errors],
                    status_code=400,
                )

            ACTIVE_TEMPLATE_DOCX_PATH.parent.mkdir(parents=True, exist_ok=True)
            ACTIVE_TEMPLATE_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)

            shutil.copy2(candidate, ACTIVE_TEMPLATE_DOCX_PATH)
            metadata.write_metadata(ACTIVE_TEMPLATE_JSON_PATH, template_payload)
    except (OSError, metadata.zipfile.BadZipFile, metadata.ElementTree.ParseError) as error:
        return _render_admin_dashboard(
            template_upload_errors=[
                "Template upload failed while reading the DOCX.",
                str(error),
            ],
            status_code=400,
        )

    flash(
        f"Template uploaded: {uploaded.filename}. "
        f"Saved as {ACTIVE_TEMPLATE_DOCX_PATH.name} and generated {ACTIVE_TEMPLATE_JSON_PATH.name}."
    )
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/submission/<int:submission_id>/marking-eligibility", methods=["POST"])
@login_required
def admin_update_marking_eligibility(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    action = request.form.get("action", "").strip().lower()
    reason = request.form.get("reason", "").strip()

    if action == "exclude":
        db.update_marking_exclusion(submission_id, True, reason=reason)
        flash(f"Submission #{submission_id} excluded from marking.")
    elif action == "include":
        db.update_marking_exclusion(submission_id, False)
        flash(f"Submission #{submission_id} included in marking.")
    else:
        flash("Unknown marking eligibility action.")

    search = request.form.get("q", "").strip()
    marking_filter = request.form.get("marking", "all").strip().lower()
    if marking_filter not in {"all", "included", "excluded"}:
        marking_filter = "all"
    sort_key = request.form.get("sort", "submitted_at").strip().lower()
    if sort_key not in {
        "submitted_at",
        "name",
        "student_id",
        "code",
        "status",
        "marking",
        "storage_backend",
        "email_sent",
        "ip_address",
    }:
        sort_key = "submitted_at"
    sort_dir = request.form.get("dir", "desc").strip().lower()
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc"
    return redirect(url_for("admin_dashboard", q=search, marking=marking_filter, sort=sort_key, dir=sort_dir))


@app.route("/admin/export/marking.xlsx")
@login_required
def admin_export_marking_overview_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Excel export requires openpyxl. Run: pip install -r requirements.txt")
        return redirect(url_for("admin_dashboard"))

    include_excluded = request.args.get("include_excluded", "0").strip().lower() in {"1", "true", "yes"}
    submissions = db.list_submissions(marking_filter="all" if include_excluded else "included")
    question_ids = set()
    rows = []

    for summary in submissions:
        submission = db.get_submission(summary["id"])
        if not submission:
            continue

        previews = _load_marking_previews(submission)
        for preview in previews:
            for answer in preview["data"].get("answers", []):
                qid = str(answer.get("question_id", "")).strip()
                if qid:
                    question_ids.add(qid)

        assessments = {
            item["question_id"]: item
            for item in db.list_marking_assessments(submission["id"])
            if item.get("question_id")
        }

        # Keep any manually saved assessments even when a preview question is missing.
        for qid in assessments:
            question_ids.add(qid)

        rows.append({
            "submission": submission,
            "assessments": assessments,
        })

    def question_sort_key(question_id):
        parts = re.split(r"(\d+)", question_id or "")
        normalized = []
        for part in parts:
            if part.isdigit():
                normalized.append((0, int(part)))
            else:
                normalized.append((1, part.lower()))
        return normalized

    ordered_questions = sorted(question_ids, key=question_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marking Overview"

    headers = [
        "Submission ID",
        "Submitted At",
        "Name",
        "Student ID",
        "Email",
        "Code",
        "Status",
    ]
    for qid in ordered_questions:
        headers.append(f"{qid} Score")
        headers.append(f"{qid} Feedback")
        headers.append(f"{qid} Source")
        headers.append(f"{qid} AI Reasoning")

    ws.append(headers)

    for row_data in rows:
        submission = row_data["submission"]
        assessments = row_data["assessments"]
        row = [
            submission.get("id", ""),
            submission.get("submitted_at", ""),
            submission.get("name", ""),
            submission.get("student_id", ""),
            submission.get("email", ""),
            submission.get("code", ""),
            submission.get("status", ""),
        ]
        for qid in ordered_questions:
            assessment = assessments.get(qid, {})
            row.append(assessment.get("score", ""))
            row.append(assessment.get("comment", ""))
            row.append(assessment.get("comment_source", "human"))
            row.append(assessment.get("ai_reasoning", ""))
        ws.append(row)

    ws.freeze_panes = "A2"
    for col in ws[1]:
        col.font = Font(bold=True)

    # Keep columns readable without creating very wide sheets.
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        width = min(max(12, max_len + 2), 60)
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="marking_overview.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/export/marking_students_columns.xlsx")
@login_required
def admin_export_marking_students_columns_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Excel export requires openpyxl. Run: pip install -r requirements.txt")
        return redirect(url_for("admin_dashboard"))

    submissions = db.list_submissions(marking_filter="all")
    question_ids = set()
    rows = []

    for summary in submissions:
        submission = db.get_submission(summary["id"])
        if not submission:
            continue

        previews = _load_marking_previews(submission)
        for preview in previews:
            for answer in preview["data"].get("answers", []):
                qid = str(answer.get("question_id", "")).strip()
                if qid:
                    question_ids.add(qid)

        assessments = {
            item["question_id"]: item
            for item in db.list_marking_assessments(submission["id"])
            if item.get("question_id")
        }

        # Keep any manually saved assessments even when a preview question is missing.
        for qid in assessments:
            question_ids.add(qid)

        rows.append({
            "submission": submission,
            "assessments": assessments,
        })

    def question_sort_key(question_id):
        parts = re.split(r"(\d+)", question_id or "")
        normalized = []
        for part in parts:
            if part.isdigit():
                normalized.append((0, int(part)))
            else:
                normalized.append((1, part.lower()))
        return normalized

    ordered_questions = sorted(question_ids, key=question_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students by Question"

    student_headers = [
        f"{item['submission'].get('student_id', '')} | {item['submission'].get('name', '')}"
        for item in rows
    ]

    headers = ["Question", "Field", *student_headers]
    ws.append(headers)
    ws.freeze_panes = "C2"

    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_fields = [
        ("Score", "score"),
        ("Feedback", "comment"),
        ("Source", "comment_source"),
        ("AI Reasoning", "ai_reasoning"),
    ]

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    last_col = max(2, len(headers))

    for qid in ordered_questions:
        block_start = ws.max_row + 1

        for idx, (label, key) in enumerate(row_fields):
            row = [qid if idx == 0 else "", label]
            for row_data in rows:
                assessment = row_data["assessments"].get(qid, {})
                value = assessment.get(key, "")
                if key == "comment_source" and value == "":
                    value = ""
                row.append(value)
            ws.append(row)

        block_end = ws.max_row

        for r in range(block_start, block_end + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=medium if c == 1 else thin,
                    right=medium if c == last_col else thin,
                    top=medium if r == block_start else thin,
                    bottom=medium if r == block_end else thin,
                )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14

    for idx in range(3, last_col + 1):
        letter = get_column_letter(idx)
        max_len = 0
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=idx, max_col=idx):
            value = "" if row_cells[0].value is None else str(row_cells[0].value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[letter].width = min(max(22, max_len + 2), 48)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="marking_students_columns.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/submissions/delete-all", methods=["POST"])
@login_required
def admin_delete_all_submissions():
    confirmation = request.form.get("confirmation", "").strip()
    if confirmation != "DELETE ALL":
        flash("Delete-all cancelled: type DELETE ALL to confirm.")
        return redirect(url_for("admin_dashboard"))

    upload_root = Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"]).resolve()
    db.delete_all_submissions(reset_ids=True)

    if upload_root.exists():
        for child in upload_root.iterdir():
            try:
                if child.is_dir():
                    shutil.rmtree(child)
                else:
                    child.unlink()
            except OSError:
                continue

    flash("All submissions were deleted and numbering was reset.")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/submission/<int:submission_id>")
@login_required
def admin_submission_detail(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)
    return render_template(
        "admin_detail.html",
        submission=submission,
        form_version=SUBMISSION_CONFIG.get("form_version"),
    )


@app.route("/admin/submission/<int:submission_id>/analysis", methods=["GET", "POST"])
@login_required
def admin_submission_analysis(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    if request.method == "POST":
        run_files = []
        seen_paths = set()
        for file in submission["files"]:
            path = Path(file["storage_location"])
            if path.suffix.lower() == ".run" and not path.name.startswith(".") and path.exists():
                resolved = path.resolve()
                if resolved not in seen_paths:
                    run_files.append(resolved)
                    seen_paths.add(resolved)

        if not run_files:
            flash("No runnable .run files were found in this submission.")
            return redirect(url_for("admin_submission_analysis", submission_id=submission_id))

        timeout_seconds = int(os.environ.get("AMPL_RUN_TIMEOUT_SECONDS", "120"))
        runner = Path(__file__).parent / "src" / "analysis_runner.py"
        ampl_python = os.environ.get("AMPL_PYTHON", sys.executable)
        submission_root = run_files[0].parents[1]
        result_root = submission_root / "analysis"
        for run_path in run_files:
            result_directory = result_root / secure_filename(run_path.stem)
            result_directory.mkdir(parents=True, exist_ok=True)
            run_id = db.create_analysis_run(
                submission_id, run_path.name, str(result_directory)
            )
            result = {}
            try:
                completed = subprocess.run(
                    [ampl_python, str(runner), str(run_path), str(run_path.parent), str(result_directory)],
                    capture_output=True,
                    text=True,
                    timeout=timeout_seconds,
                    check=False,
                )
                try:
                    result = json.loads(completed.stdout.strip().splitlines()[-1])
                except (json.JSONDecodeError, IndexError):
                    result = {
                        "stderr": completed.stderr,
                        "error": "The AMPL worker returned an invalid result.",
                        "statistics": {},
                    }
                status = "completed" if result.get("status") == "completed" else "failed"
            except subprocess.TimeoutExpired as error:
                timeout_stdout = error.stdout or ""
                timeout_stderr = error.stderr or ""
                if isinstance(timeout_stdout, bytes):
                    timeout_stdout = timeout_stdout.decode(errors="replace")
                if isinstance(timeout_stderr, bytes):
                    timeout_stderr = timeout_stderr.decode(errors="replace")
                result = {
                    "error": f"AMPL execution exceeded the {timeout_seconds}-second timeout.",
                    "stdout": timeout_stdout,
                    "stderr": timeout_stderr,
                    "statistics": {},
                }
                status = "timed_out"
            except OSError as error:
                result = {
                    "error": f"Could not start AMPL Python interpreter '{ampl_python}'.",
                    "stderr": str(error),
                    "statistics": {},
                    "diagnostics": {"python_executable": ampl_python},
                }
                status = "failed"
            db.update_analysis_run(run_id, status, result)

        flash(f"AMPL analysis completed for {len(run_files)} run file(s).")
        return redirect(url_for("admin_submission_analysis", submission_id=submission_id))

    analysis_text_files_by_run = {
        run["id"]: _collect_analysis_text_files_for_run(submission, run)
        for run in submission["analysis_runs"]
    }

    return render_template(
        "admin_analysis.html",
        submission=submission,
        analysis_text_files_by_run=analysis_text_files_by_run,
        form_version=SUBMISSION_CONFIG.get("form_version"),
    )


@app.route("/admin/submission/<int:submission_id>/analysis/<int:run_id>/delete", methods=["POST"])
@login_required
def admin_delete_analysis_run(submission_id, run_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    run = next((item for item in submission["analysis_runs"] if item["id"] == run_id), None)
    if not run:
        abort(404)

    result_location = Path(run["result_location"]).resolve()
    submission_root = Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"]).resolve()
    db.delete_analysis_run(run_id, submission_id)
    if submission_root in result_location.parents:
        if result_location.is_dir():
            shutil.rmtree(result_location)
        elif result_location.exists():
            result_location.unlink()

    flash(f"Analysis run '{run['run_filename']}' was deleted.")
    return redirect(url_for("admin_submission_analysis", submission_id=submission_id))


@app.route("/admin/submission/<int:submission_id>/analysis/<int:run_id>/text-file")
@login_required
def admin_analysis_text_file(submission_id, run_id):
    submission = db.get_submission(submission_id)
    if not submission:
        return jsonify({"ok": False, "error": "Submission not found."}), 404

    run_record = next((item for item in submission.get("analysis_runs", []) if item.get("id") == run_id), None)
    if not run_record:
        return jsonify({"ok": False, "error": "Analysis run not found."}), 404

    relative_path = request.args.get("path", "").strip()
    if not relative_path:
        return jsonify({"ok": False, "error": "Missing file path."}), 400

    allowed_files = _collect_analysis_text_files_for_run(submission, run_record)
    allowed_paths = {item["path"] for item in allowed_files}
    if relative_path not in allowed_paths:
        return jsonify({"ok": False, "error": "Requested file is not available for this run."}), 404

    submission_root = _submission_root_folder(submission)
    if not submission_root:
        return jsonify({"ok": False, "error": "Submission files are not available."}), 404

    try:
        resolved_root = submission_root.resolve()
        candidate = (resolved_root / relative_path).resolve()
        candidate.relative_to(resolved_root)
    except (OSError, ValueError):
        return jsonify({"ok": False, "error": "Invalid file path."}), 400

    if not candidate.exists() or not candidate.is_file():
        return jsonify({"ok": False, "error": "File does not exist."}), 404

    if candidate.suffix.lower() not in MARKING_TEXT_PREVIEW_EXTENSIONS:
        return jsonify({"ok": False, "error": "Only .mod, .run, and .dat files can be previewed."}), 400

    try:
        raw = candidate.read_bytes()
    except OSError as error:
        return jsonify({"ok": False, "error": f"Could not read file: {error}"}), 500

    truncated = len(raw) > MARKING_TEXT_PREVIEW_MAX_BYTES
    if truncated:
        raw = raw[:MARKING_TEXT_PREVIEW_MAX_BYTES]

    content = raw.decode("utf-8", errors="replace")
    return jsonify({
        "ok": True,
        "file": {
            "path": relative_path,
            "name": candidate.name,
            "content": content,
            "truncated": truncated,
            "max_bytes": MARKING_TEXT_PREVIEW_MAX_BYTES,
        },
    })


@app.route("/admin/submission/<int:submission_id>/metadata")
@login_required
def admin_submission_metadata(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    metadata_entries = []
    local_files = [Path(file["storage_location"]) for file in submission["files"]]
    if local_files:
        metadata_folder = local_files[0].parents[1] / "metadata"
        for metadata_path in sorted(metadata_folder.glob("*.json")):
            try:
                if metadata_path.name.endswith(".answers.json"):
                    # Generated extraction preview, not a student-submitted file.
                    continue
                data = json.loads(metadata_path.read_text(encoding="utf-8"))
                if data.get("extracted_from"):
                    # Metadata for files unpacked from a submitted archive.
                    continue
                original_name = data.get("original_filename", "")
                if not original_name:
                    # Defensive guard: skip JSON files that are not file-metadata records.
                    continue
                display_name = Path(original_name).stem if original_name else "Metadata entry"
                metadata_entries.append({
                    "filename": data.get("original_filename", metadata_path.name),
                    "display_name": display_name,
                    "data": data,
                })
            except (OSError, json.JSONDecodeError):
                continue

    return render_template(
        "admin_metadata.html", submission=submission, metadata_entries=metadata_entries,
        form_version=SUBMISSION_CONFIG.get("form_version"),
    )


@app.route("/admin/submission/<int:submission_id>/marking")
@login_required
def admin_submission_marking(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    view_mode = request.args.get("view", "student").strip().lower()
    if view_mode not in {"student", "question"}:
        view_mode = "student"
    show_only_unmarked = request.args.get("unmarked", "0").strip().lower() in {"1", "true", "yes"}

    def assessment_has_marking(assessment_record):
        if not assessment_record:
            return False
        score = str(assessment_record.get("score", "") or "").strip()
        comment = str(assessment_record.get("comment", "") or "").strip()
        return bool(score and comment)

    active_template_context = _load_active_template_questions()
    template_questions = active_template_context["questions"]
    template_question_ids = active_template_context["question_ids"]
    template_question_map = {item["question_id"]: item for item in template_questions}

    def extract_answers_by_question(previews):
        ordered = []
        seen = set()
        by_question = {}
        for preview in previews:
            for answer in preview["data"].get("answers", []):
                question_id = _normalize_question_id(answer.get("question_id"))
                if not question_id:
                    continue

                answer_copy = dict(answer)
                answer_copy["question_id"] = question_id
                answer_copy["template_question_id"] = (
                    _normalize_question_id(answer_copy.get("template_question_id")) or question_id
                )

                if question_id not in seen:
                    seen.add(question_id)
                    ordered.append(question_id)

                if question_id not in by_question:
                    by_question[question_id] = {
                        "preview_file": preview["file"],
                        "answer": answer_copy,
                        "images": answer_copy.get("images", []),
                    }

        return ordered, by_question

    def merge_question_ids(*question_id_lists):
        merged = []
        seen = set()
        extras = []
        for index, values in enumerate(question_id_lists):
            for raw in values or []:
                question_id = _normalize_question_id(raw)
                if not question_id or question_id in seen:
                    continue
                seen.add(question_id)
                if index == 0:
                    merged.append(question_id)
                else:
                    extras.append(question_id)

        for question_id in sorted(set(extras), key=_question_sort_key):
            if question_id not in merged:
                merged.append(question_id)
        return merged

    def build_answer_entry(question_id, matched):
        template_meta = template_question_map.get(question_id, {})
        template_prompt = str(template_meta.get("question_prompt", "") or "").strip()
        template_marks_label = str(template_meta.get("marks_label", "") or "").strip()

        source_answer = dict((matched or {}).get("answer") or {})
        extracted_prompt = str(source_answer.get("prompt", "") or "").strip()
        merged_prompt = extracted_prompt or template_prompt

        answer_entry = {
            "question_id": question_id,
            "template_question_id": source_answer.get("template_question_id") or question_id,
            "prompt": merged_prompt,
            "template_prompt": template_prompt,
            "extracted_prompt": extracted_prompt,
            "marks_label": source_answer.get("marks_label") or template_marks_label,
            "answer": str(source_answer.get("answer", "") or ""),
            "answer_paragraphs": source_answer.get("answer_paragraphs") or [],
            "confidence": source_answer.get("confidence"),
            "images": source_answer.get("images") or [],
            "max_score": template_meta.get("max_score"),
        }

        return {
            "preview_file": (matched or {}).get("preview_file"),
            "answer": answer_entry,
            "images": answer_entry["images"],
        }

    def has_unmarked_items(previews, assessment_map):
        extracted_ids, _ = extract_answers_by_question(previews)
        question_ids = merge_question_ids(template_question_ids, extracted_ids, assessment_map.keys())
        for question_id in question_ids:
            if not assessment_has_marking(assessment_map.get(question_id, {})):
                return True
        return False

    previews = _load_marking_previews(submission)
    extracted_question_ids, answers_by_question = extract_answers_by_question(previews)
    assessments = {
        item["question_id"]: item
        for item in db.list_marking_assessments(submission_id)
    }
    question_ids = merge_question_ids(template_question_ids, extracted_question_ids, assessments.keys())

    student_question_entries = []
    for question_id in question_ids:
        matched = answers_by_question.get(question_id)
        answer_block = build_answer_entry(question_id, matched)
        student_question_entries.append({
            "question_id": question_id,
            "assessment": assessments.get(question_id, {}),
            "eval_candidate": db.get_eval_case_candidate(submission_id, question_id) or {},
            **answer_block,
        })

    can_reextract = not db.has_marking_assessments(submission_id)

    included_submissions = db.list_submissions(marking_filter="included")
    submission_ids = [item["id"] for item in included_submissions]
    submission_is_excluded = bool(submission.get("marking_excluded"))

    unmarked_submission_ids = []
    if show_only_unmarked:
        for item in included_submissions:
            full_submission = db.get_submission(item["id"])
            if not full_submission:
                continue
            full_previews = _load_marking_previews(full_submission)
            full_assessments = {
                assessment["question_id"]: assessment
                for assessment in db.list_marking_assessments(full_submission["id"])
            }
            if has_unmarked_items(full_previews, full_assessments):
                unmarked_submission_ids.append(full_submission["id"])

    if view_mode == "student" and show_only_unmarked and unmarked_submission_ids and submission_id not in unmarked_submission_ids:
        return redirect(
            url_for(
                "admin_submission_marking",
                submission_id=unmarked_submission_ids[0],
                view="student",
                unmarked="1",
            )
        )

    nav_submission_ids = unmarked_submission_ids if show_only_unmarked else submission_ids
    prev_submission_id = None
    next_submission_id = None
    if submission_id in nav_submission_ids:
        current_index = nav_submission_ids.index(submission_id)
        if current_index > 0:
            prev_submission_id = nav_submission_ids[current_index - 1]
        if current_index + 1 < len(nav_submission_ids):
            next_submission_id = nav_submission_ids[current_index + 1]

    student_all_marked = bool(show_only_unmarked and not unmarked_submission_ids)

    selected_question = request.args.get("question", "").strip()
    question_entries = []
    question_prev_submission_id = None
    question_next_submission_id = None
    question_entry = None
    question_all_marked = False

    if view_mode == "question":
        # Build answer entries for every submission that has this question.
        all_question_ids = list(template_question_ids)
        all_seen = set()
        for question_id in template_question_ids:
            all_seen.add(question_id)
        question_has_unmarked = {}
        staged = []
        for summary in included_submissions:
            full_submission = db.get_submission(summary["id"])
            if not full_submission:
                continue
            full_previews = _load_marking_previews(full_submission)
            full_assessments = {
                item["question_id"]: item
                for item in db.list_marking_assessments(full_submission["id"])
            }

            full_extracted_ids, full_answers_by_question = extract_answers_by_question(full_previews)
            full_question_ids = merge_question_ids(
                template_question_ids,
                full_extracted_ids,
                full_assessments.keys(),
            )

            for qid in full_question_ids:
                if qid not in all_seen:
                    all_seen.add(qid)
                    all_question_ids.append(qid)
                if not assessment_has_marking(full_assessments.get(qid, {})):
                    question_has_unmarked[qid] = True

            staged.append((full_submission, full_assessments, full_answers_by_question, full_question_ids))

        if show_only_unmarked:
            all_question_ids = [qid for qid in all_question_ids if question_has_unmarked.get(qid)]

        if not selected_question and all_question_ids:
            selected_question = all_question_ids[0]
        if selected_question and selected_question not in all_question_ids:
            selected_question = all_question_ids[0] if all_question_ids else ""

        question_ids = all_question_ids
        question_all_marked = bool(show_only_unmarked and not question_ids)

        if selected_question:
            for full_submission, full_assessments, full_answers_by_question, full_question_ids in staged:
                if selected_question not in full_question_ids:
                    continue
                selected_assessment = full_assessments.get(selected_question, {})
                if show_only_unmarked and assessment_has_marking(selected_assessment):
                    continue
                eval_candidate = db.get_eval_case_candidate(full_submission["id"], selected_question)
                answer_block = build_answer_entry(
                    selected_question,
                    full_answers_by_question.get(selected_question),
                )
                question_entries.append({
                    "submission_id": full_submission["id"],
                    "name": full_submission["name"],
                    "student_id": full_submission["student_id"],
                    "email": full_submission.get("email"),
                    "submitted_at": full_submission["submitted_at"],
                    "assessment": selected_assessment,
                    "eval_candidate": eval_candidate,
                    **answer_block,
                })

        if question_entries:
            entry_ids = [entry["submission_id"] for entry in question_entries]
            if submission_id in entry_ids:
                active_index = entry_ids.index(submission_id)
            else:
                active_index = 0
            question_entry = question_entries[active_index]
            if active_index > 0:
                question_prev_submission_id = question_entries[active_index - 1]["submission_id"]
            if active_index + 1 < len(question_entries):
                question_next_submission_id = question_entries[active_index + 1]["submission_id"]

    marking_text_submission = submission
    if view_mode == "question" and question_entry and question_entry["submission_id"] != submission["id"]:
        marking_text_submission = db.get_submission(question_entry["submission_id"]) or submission

    active_question_for_files = selected_question if view_mode == "question" else None
    marking_text_files = _collect_marking_text_files(marking_text_submission, question_id=active_question_for_files)
    marking_document_files = _collect_marking_document_files(marking_text_submission, question_id=active_question_for_files)
    marking_document_options = []

    for item in marking_document_files:
        extension = item.get("extension", "")
        if item.get("media_type") == "video":
            marking_document_options.append({
                "path": item["path"],
                "mode": "video",
                "label": f"{item['label']} [Video]",
                "priority": 0,
            })
        elif extension == ".pdf":
            marking_document_options.append({
                "path": item["path"],
                "mode": "pdf",
                "label": f"{item['label']} [PDF]",
                "priority": 1,
            })
        elif extension == ".docx":
            marking_document_options.append({
                "path": item["path"],
                "mode": "converted_pdf",
                "label": f"{item['label']} [PDF preview]",
                "priority": 2,
            })
            marking_document_options.append({
                "path": item["path"],
                "mode": "docx",
                "label": f"{item['label']} [DOCX view]",
                "priority": 3,
            })

    marking_document_options.sort(key=lambda item: (item["priority"], item["label"].lower()))

    default_option = None
    for preferred_mode in ("pdf", "converted_pdf", "docx"):
        default_option = next((item for item in marking_document_options if item.get("mode") == preferred_mode), None)
        if default_option:
            break
    if not default_option and marking_document_options:
        default_option = marking_document_options[0]
    for item in marking_document_options:
        item["selected"] = item is default_option

    return render_template(
        "admin_marking.html",
        submission=submission,
        submission_is_excluded=submission_is_excluded,
        view_mode=view_mode,
        previews=previews,
        question_ids=question_ids,
        selected_question=selected_question,
        student_question_entries=student_question_entries,
        question_entry=question_entry,
        assessments=assessments,
        can_reextract=can_reextract,
        active_template_status=active_template_context,
        prev_submission_id=prev_submission_id,
        next_submission_id=next_submission_id,
        question_prev_submission_id=question_prev_submission_id,
        question_next_submission_id=question_next_submission_id,
        show_only_unmarked=show_only_unmarked,
        student_all_marked=student_all_marked,
        question_all_marked=question_all_marked,
        marking_text_files=marking_text_files,
        marking_document_options=marking_document_options,
        marking_text_submission_id=marking_text_submission["id"],
        form_version=SUBMISSION_CONFIG.get("form_version"),
    )


@app.route("/admin/submission/<int:submission_id>/marking/text-file")
@login_required
def admin_marking_text_file(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        return jsonify({"ok": False, "error": "Submission not found."}), 404

    relative_path = request.args.get("path", "").strip()
    if not relative_path:
        return jsonify({"ok": False, "error": "Missing file path."}), 400

    text_file_entries = _collect_marking_text_files(submission)
    allowed_paths = {entry["path"] for entry in text_file_entries}
    if relative_path not in allowed_paths:
        return jsonify({"ok": False, "error": "Requested file is not available for preview."}), 404

    submission_root = _submission_root_folder(submission)
    if not submission_root:
        return jsonify({"ok": False, "error": "Submission files are not available."}), 404

    try:
        resolved_root = submission_root.resolve()
        candidate = (resolved_root / relative_path).resolve()
        candidate.relative_to(resolved_root)
    except (OSError, ValueError):
        return jsonify({"ok": False, "error": "Invalid file path."}), 400

    if not candidate.exists() or not candidate.is_file():
        return jsonify({"ok": False, "error": "File does not exist."}), 404

    if candidate.suffix.lower() not in MARKING_TEXT_PREVIEW_EXTENSIONS:
        return jsonify({"ok": False, "error": "Only .mod, .run, and .dat files can be previewed."}), 400

    try:
        raw = candidate.read_bytes()
    except OSError as error:
        return jsonify({"ok": False, "error": f"Could not read file: {error}"}), 500

    truncated = len(raw) > MARKING_TEXT_PREVIEW_MAX_BYTES
    if truncated:
        raw = raw[:MARKING_TEXT_PREVIEW_MAX_BYTES]

    content = raw.decode("utf-8", errors="replace")
    return jsonify({
        "ok": True,
        "file": {
            "path": relative_path,
            "name": candidate.name,
            "content": content,
            "truncated": truncated,
            "max_bytes": MARKING_TEXT_PREVIEW_MAX_BYTES,
        },
    })


@app.route("/admin/submission/<int:submission_id>/marking/document")
@login_required
def admin_marking_document_preview(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    relative_path = request.args.get("path", "").strip()
    preview_mode = request.args.get("mode", "").strip().lower()
    if not relative_path:
        abort(400, "Missing file path.")

    document_entries = _collect_marking_document_files(submission)
    allowed_paths = {entry["path"] for entry in document_entries}
    if relative_path not in allowed_paths:
        abort(404, "Requested file is not available for preview.")

    candidate = _resolve_submission_relative_file(submission, relative_path)
    if not candidate:
        abort(404, "File does not exist.")

    extension = candidate.suffix.lower()
    if extension in MARKING_VIDEO_PREVIEW_EXTENSIONS:
        if preview_mode in {"", "video"}:
            media_url = url_for("admin_marking_document_media", submission_id=submission_id, path=relative_path)
            media_type = mimetypes.guess_type(candidate.name)[0] or "video/mp4"
            return render_template(
                "admin_marking_media_preview.html",
                file_name=candidate.name,
                file_path=relative_path,
                media_url=media_url,
                media_type=media_type,
                media_kind="video",
                error_message="",
            )

        abort(400, "Unknown preview mode for video files.")

    if extension == ".pdf":
        return send_file(candidate, mimetype="application/pdf", as_attachment=False, download_name=candidate.name)

    if extension == ".docx":
        if preview_mode in {"", "pdf", "converted_pdf"}:
            try:
                preview_pdf = _ensure_docx_pdf_preview(submission, relative_path, candidate)
                return send_file(preview_pdf, mimetype="application/pdf", as_attachment=False, download_name=f"{candidate.stem}.pdf")
            except (OSError, RuntimeError, ValueError) as error:
                # Fall through to docx readable view with an explanatory message.
                docx_fallback_error = (
                    "PDF preview is unavailable for this DOCX in the current environment. "
                    f"Showing DOCX view instead. ({escape(str(error))})"
                )
        else:
            docx_fallback_error = ""

        try:
            paragraphs = metadata._docx_paragraphs(candidate)
            images = metadata._docx_images(candidate, max_images=18, max_total_bytes=10 * 1024 * 1024)
        except (OSError, KeyError, metadata.zipfile.BadZipFile, metadata.ElementTree.ParseError) as error:
            return render_template(
                "admin_marking_docx_preview.html",
                file_name=candidate.name,
                file_path=relative_path,
                paragraphs=[],
                images=[],
                truncated=False,
                error_message=f"Could not render DOCX preview: {escape(str(error))}",
            )

        truncated = len(paragraphs) > MARKING_DOCX_PREVIEW_MAX_PARAGRAPHS
        if truncated:
            paragraphs = paragraphs[:MARKING_DOCX_PREVIEW_MAX_PARAGRAPHS]

        return render_template(
            "admin_marking_docx_preview.html",
            file_name=candidate.name,
            file_path=relative_path,
            paragraphs=paragraphs,
            images=images,
            truncated=truncated,
            max_paragraphs=MARKING_DOCX_PREVIEW_MAX_PARAGRAPHS,
            error_message=docx_fallback_error,
        )

    abort(400, "Only PDF and DOCX files are supported.")


@app.route("/admin/submission/<int:submission_id>/marking/media")
@login_required
def admin_marking_document_media(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    relative_path = request.args.get("path", "").strip()
    if not relative_path:
        abort(400, "Missing file path.")

    document_entries = _collect_marking_document_files(submission)
    allowed_paths = {entry["path"] for entry in document_entries}
    if relative_path not in allowed_paths:
        abort(404, "Requested file is not available for preview.")

    candidate = _resolve_submission_relative_file(submission, relative_path)
    if not candidate:
        abort(404, "File does not exist.")

    extension = candidate.suffix.lower()
    if extension not in MARKING_VIDEO_PREVIEW_EXTENSIONS:
        abort(400, "Only video files are supported.")

    mimetype = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
    return send_file(candidate, mimetype=mimetype, as_attachment=False, download_name=candidate.name)


@app.route("/admin/submission/<int:submission_id>/marking/assessment", methods=["POST"])
@login_required
def admin_save_marking_assessment(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    question_id = request.form.get("question_id", "").strip()
    score = request.form.get("score", "").strip()
    comment = request.form.get("comment", "").strip()
    view_mode = request.form.get("view", "student").strip().lower()
    selected_question = request.form.get("selected_question", "").strip()
    unmarked_only = request.form.get("unmarked", "0").strip().lower() in {"1", "true", "yes"}
    question_prompt = request.form.get("question_prompt", "").strip()
    student_answer = request.form.get("student_answer", "").strip()
    marks_label = request.form.get("marks_label", "").strip()
    include_eval_case = request.form.get("include_eval_case") == "true"
    ai_draft_used = request.form.get("ai_draft_used") == "true"
    comment_source = request.form.get("comment_source", "human").strip().lower()
    ai_reasoning = request.form.get("ai_reasoning", "").strip()
    is_async = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if comment_source not in {"ai", "ai_human_approved", "human"}:
        comment_source = "human"
    if comment_source == "human":
        ai_reasoning = ""

    if not question_id:
        if is_async:
            return jsonify({"ok": False, "error": "Question ID is required for saving marking."}), 400
        flash("Question ID is required for saving marking.")
    else:
        db.save_marking_assessment(
            submission_id,
            question_id,
            score,
            comment,
            comment_source=comment_source,
            ai_reasoning=ai_reasoning,
        )
        message = f"Saved marking for {question_id}."
        detail = ""

        # Auto-capture manual (non-AI) entries into evaluation candidates.
        if comment_source == "human" and not ai_draft_used:
            if question_prompt and student_answer and score and comment:
                db.save_eval_case_candidate(
                    submission_id=submission_id,
                    question_id=question_id,
                    question_prompt=question_prompt,
                    student_answer=student_answer,
                    marks_label=marks_label,
                    reference_score=score,
                    reference_comment=comment,
                    ai_draft_used=False,
                    include_in_eval=include_eval_case,
                )
            else:
                detail = (
                    f"Evaluation candidate for {question_id} was skipped "
                    "(missing prompt/answer/mark fields)."
                )
        else:
            detail = (
                f"{question_id} was saved as AI-assisted marking and not auto-added "
                "as an evaluation candidate."
            )

        if is_async:
            return jsonify({
                "ok": True,
                "message": message,
                "detail": detail,
            })

        flash(message)
        if detail:
            flash(detail)

    return redirect(
        url_for(
            "admin_submission_marking",
            submission_id=submission_id,
            view=view_mode if view_mode in {"student", "question"} else "student",
            question=selected_question or None,
            unmarked="1" if unmarked_only else None,
        )
    )


@app.route("/admin/submission/<int:submission_id>/marking/assessment/undo", methods=["POST"])
@login_required
def admin_undo_marking_assessment(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    question_id = request.form.get("question_id", "").strip()
    view_mode = request.form.get("view", "student").strip().lower()
    selected_question = request.form.get("selected_question", "").strip()
    unmarked_only = request.form.get("unmarked", "0").strip().lower() in {"1", "true", "yes"}

    if not question_id:
        flash("Question ID is required for undoing marking.")
    else:
        db.delete_marking_assessment(submission_id, question_id)
        db.delete_eval_case_candidate(submission_id, question_id)
        flash(f"Undid marking for {question_id}.")

    return redirect(
        url_for(
            "admin_submission_marking",
            submission_id=submission_id,
            view=view_mode if view_mode in {"student", "question"} else "student",
            question=selected_question or None,
            unmarked="1" if unmarked_only else None,
        )
    )


@app.route("/admin/submission/<int:submission_id>/marking/ai-suggest", methods=["POST"])
@login_required
def admin_marking_ai_suggest(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        return jsonify({"ok": False, "error": "Submission not found."}), 404

    payload = request.get_json(silent=True) or {}
    question_id = str(payload.get("question_id", "")).strip()
    prompt = str(payload.get("question_prompt", "")).strip()
    answer = str(payload.get("student_answer", "")).strip()
    marks_label = str(payload.get("marks_label", "")).strip()

    if not question_id or not prompt or not answer:
        return jsonify({
            "ok": False,
            "error": "question_id, question_prompt, and student_answer are required.",
        }), 400

    try:
        result = _generate_ai_marking_suggestion(question_id, prompt, answer, marks_label)
    except urlerror.URLError as error:
        return jsonify({
            "ok": False,
            "error": f"Could not connect to Ollama at {OLLAMA_ENDPOINT}: {error}",
        }), 503
    except (ValueError, json.JSONDecodeError) as error:
        return jsonify({
            "ok": False,
            "error": f"Invalid response from Ollama: {error}",
        }), 502

    if not result.get("ok"):
        return jsonify(result), 502
    return jsonify(result)


@app.route("/admin/submission/<int:submission_id>/marking/reextract", methods=["POST"])
@login_required
def admin_reextract_marking_preview(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    if db.has_marking_assessments(submission_id):
        flash("Re-extract is blocked because marking scores/comments already exist for this submission.")
        return redirect(url_for("admin_submission_marking", submission_id=submission_id, view="student"))

    extraction_areas = set(SUBMISSION_CONFIG.get("marking_extraction_areas", ["report"]))
    updated = 0
    errors = 0

    for file_record in submission["files"]:
        if file_record["area_key"] not in extraction_areas:
            continue
        try:
            if _extract_marking_preview_for_submission_file(submission, file_record):
                updated += 1
        except (OSError, metadata.zipfile.BadZipFile, metadata.ElementTree.ParseError):
            errors += 1

    active_template_note = ""
    if ACTIVE_TEMPLATE_DOCX_PATH.exists():
        active_template_note = f" using {ACTIVE_TEMPLATE_DOCX_PATH.name}"

    if updated:
        flash(f"Re-extracted marking preview for {updated} file(s){active_template_note}.")
    elif errors:
        flash("Re-extract failed for all candidate files.")
    else:
        flash("No eligible local DOCX files found for re-extraction.")

    return redirect(url_for("admin_submission_marking", submission_id=submission_id, view="student"))


@app.route("/admin/submission/<int:submission_id>/delete", methods=["POST"])
@login_required
def admin_delete_submission(submission_id):
    submission = db.get_submission(submission_id)
    if not submission:
        abort(404)

    local_files = [Path(file["storage_location"]) for file in submission["files"]]
    db.delete_submission(submission_id)

    if local_files:
        upload_root = Path(STORAGE_ENV["LOCAL_UPLOAD_ROOT"]).resolve()
        submission_folder = local_files[0].resolve().parents[1]
        if upload_root in submission_folder.parents:
            shutil.rmtree(submission_folder)

    flash(f"Submission #{submission_id} was deleted.")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/download/<int:file_id>")
@login_required
def admin_download_file(file_id):
    conn = db.get_connection()
    row = conn.execute("SELECT * FROM submission_files WHERE id = ?", (file_id,)).fetchone()
    conn.close()
    if not row:
        abort(404)
    row = dict(row)
    # Only local-backend files can be served directly; cloud files are opened via the provider.
    path = Path(row["storage_location"])
    if not path.exists():
        abort(404, "File not found on local storage (it may be stored in the cloud backend instead).")
    return send_file(path, as_attachment=True, download_name=row["original_filename"])


if __name__ == "__main__":
    db.init_db()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
